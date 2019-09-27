import os
import sys
import cantools
from pprint import pprint
from openpyxl import load_workbook
from openpyxl import Workbook
from unidecode import unidecode


#----------------------------------------------------------------------------
#---------------------------   VAR   ----------------------------------------
#----------------------------------------------------------------------------
DBC_PATH = "C:\\Work\\Download\\DBC\\"
DBC_FILE_NAME = "Matrice_PASS_FRONT_RADAR_G5_FD3.dbc"
ExcelSheetPath = "C:\\Work\\DOC\\"
ExcelSheetName= "Data_Conversion_Form_R5.1.xlsx"
ExcelSheet = ExcelSheetPath + ExcelSheetName
Sysvar_File_Name = "sysvarList.vsysvar"
Sysvar_File = DBC_PATH + "YJE_"+Sysvar_File_Name
CAPL_File_Name = "CAPL.capl"
CAPL_File_PATH = "C:\\Work\\Download\\DBC\\"
CAPL_File = CAPL_File_PATH + CAPL_File_Name
#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def get_Input_Signal_List_From_ExcelSheet(NameSheet):
    TabSignal=[]
    workBook = load_workbook (filename=NameSheet)
    workSheet = workBook['CAN']
    i=3
    End=False
    for x in range (3 , 220) :
        Col_Signal_Name=workSheet.cell(row = x, column=2)
        In_Out_Signal_Type=workSheet.cell(row = x, column=6)
        if (In_Out_Signal_Type.value == "In") and ("Adasis…" not in Col_Signal_Name.value):
            #In signal
            TabSignal.append(Col_Signal_Name.value)
    return TabSignal

#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def write_Head(SysvarFile):
    SysvarFile.write("<?xml version=\"1.0\" encoding=\"utf-8\"?>\n")
    SysvarFile.write("<systemvariables version=\"4\">\n")
    SysvarFile.write("  <namespace name=\"\" comment=\"\" interface=\"\">\n")
    SysvarFile.write("    <namespace name=\"RADAR_SYS_VAR\" comment=\"\" interface=\"\">\n")

#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def write_Tail(SysvarFile):
    SysvarFile.write("    </namespace>\n")
    SysvarFile.write("  </namespace>\n")
    SysvarFile.write("</systemvariables>\n")

#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def write_Line(SysvarFile,UNIT,NAME_Sig,ISSIGNED,TYPE,STARTVALUE,MINVALUE,MAXVALUE):
    SysvarFile.write("      <variable anlyzLocal=\"2\" readOnly=\"false\" valueSequence=\"false\" unit=\""+UNIT+"\" name=\""+NAME_Sig+"\" comment=\"\" bitcount=\"64\" isSigned=\""+ISSIGNED+"\" encoding=\"65001\" type=\""+TYPE+"\" startValue=\""+STARTVALUE+"\" minValue=\""+MINVALUE+"\" maxValue=\""+MAXVALUE+"\" />\n")

#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def write_Line_Value_Table(SysvarFile,UNIT,NAME_Sig,ISSIGNED,TYPE,STARTVALUE,MINVALUE,MAXVALUE):
    SysvarFile.write("      <variable anlyzLocal=\"2\" readOnly=\"false\" valueSequence=\"false\" unit=\""+UNIT+"\" name=\""+NAME_Sig+"\" comment=\"\" bitcount=\"64\" isSigned=\""+ISSIGNED+"\" encoding=\"65001\" type=\""+TYPE+"\" startValue=\""+STARTVALUE+"\" minValue=\""+MINVALUE+"\" maxValue=\""+MAXVALUE+"\">\n")

#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def DEL_Space_Accent(STR_NOK):
    STR_NOK = STR_NOK.replace(" ","_")
    STR_NOK = STR_NOK.replace("\'",".")
    STR_NOK = STR_NOK.replace("/","_")
    STR_NOK = STR_NOK.replace(":","_")
    #now we have to change e and o with accent to e and o without accent
    STR_NOK = STR_NOK.replace("Ã©","e")
    STR_NOK = STR_NOK.replace("Ã´","o")
    STR_NOK = STR_NOK.replace("Ã¨","e")
    STR_NOK = STR_NOK.replace("Ã","a")
    STR_NOK = STR_NOK.replace("Ãª","e")
    STR_NOK = STR_NOK.replace("aª","e")
    STR_OK = STR_NOK
    return STR_OK
#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def write_Value_Table(SysvarFile,SIG_VAL_TABLE):
    SysvarFile.write("        <valuetable name=\"Custom\" definesMinMax=\"false\">\n")
    List_Len = len(SIG_VAL_TABLE)
    String_List = ","
    Counter_Duplicated = 0
    for i in range(0,List_Len) :
        VAL = str(i)
        if (("," + SIG_VAL_TABLE[i]+",") in String_List ) :
            DESC = SIG_VAL_TABLE[i] + "_" +str(Counter_Duplicated)
            Counter_Duplicated = Counter_Duplicated + 1
        else :
            DESC = SIG_VAL_TABLE[i]
        String_List = String_List + SIG_VAL_TABLE[i]+","
        DESC = DEL_Space_Accent(DESC)
        SysvarFile.write("          <valuetableentry value=\""+VAL+"\" lowerBound=\""+VAL+"\" upperBound=\""+VAL+"\" description=\""+DESC+"\" displayString=\""+DESC+"\" />\n")
    SysvarFile.write("        </valuetable>\n")


#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------

def Write_List_Signal(TabSignal , CAN_Database, SysvarFile) :
    for Signal_Name_From_List in TabSignal :
        for mess in DB_database.messages :
            for signal in mess.signals :
                UNIT =""
                NAME_Sig="_Sig"
                ISSIGNED="false"
                TYPE="int"
                STARTVALUE=0
                MINVALUE=0
                MAXVALUE=0
                if (Signal_Name_From_List == signal.name) :
                    #signal found
                    if (signal.unit is None):
                        UNIT=""
                    else :
                        UNIT=signal.unit
                    NAME_Sig=signal.name + "_Sig"
                    ISSIGNED=signal.is_signed
                    if (ISSIGNED) :
                        str_ISSIGNED = "true"
                    else :
                        str_ISSIGNED = "false"

                    if (signal._is_float):
                        TYPE="float"
                    else :
                        TYPE="int"
                    STARTVALUE=signal.initial
                    MINVALUE=signal.minimum
                    MAXVALUE=signal.maximum
                    #print( UNIT, NAME_Sig ,ISSIGNED , TYPE, STARTVALUE,MINVALUE,MAXVALUE)
                    if (TYPE == "int") and (signal.choices is not None):
                        #check is table
                        write_Line_Value_Table(SysvarFile, UNIT, NAME_Sig ,str_ISSIGNED , TYPE, str(MINVALUE),str(MINVALUE),str(MAXVALUE))
                        #print("Sig name : " ,NAME_Sig)
                        write_Value_Table(SysvarFile,signal.choices)
                    else :
                        write_Line(SysvarFile, UNIT, NAME_Sig ,str_ISSIGNED , TYPE,str(MINVALUE),str(MINVALUE),str(MAXVALUE))



#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def Create_File_Of_System_Variable(TabSignal , CAN_Database, fileName):
    try:
        Sysvar = open(fileName ,'w')
    except IOError:
        print('cannot open', fileName)
        sys.exit(0)
    write_Head(Sysvar)
    Write_List_Signal(TabSignal , CAN_Database, Sysvar)
    write_Tail(Sysvar)
    Sysvar.close()
#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def get_Hash_Tab_From_Signal_List(TabSignal , CAN_Database):
    if((len(STR.split(' '))==4) and ("VAL_" in STR)) or ("TPS_CONF_MIN_RAP_MON" in STR) or ("FUP_TIMENSEC" in STR):
        return 1
    else :
        return 0

#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def Treat(LineSTR):
    if("VAL_" in LineSTR )  and (" 000 " in LineSTR) :#for LONGITUDINAL_REQUEST problem
        #problem still in database
        LineSTR = LineSTR.replace( "000" , "0")
        LineSTR = LineSTR.replace( "001" , "1")
        LineSTR = LineSTR.replace( "010" , "2")
        LineSTR = LineSTR.replace( "011" , "3")
        LineSTR = LineSTR.replace( "100" , "4")
        LineSTR = LineSTR.replace( "101" , "5")
        LineSTR = LineSTR.replace( "110" , "6")
        LineSTR = LineSTR.replace( "111" , "7")
        return LineSTR
    else :
        #no line problem
        return LineSTR

#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def check_Line_Prob(STR):
    if((len(STR.split(' '))==4) and ("VAL_" in STR)) or ("TPS_CONF_MIN_RAP_MON" in STR) or ("FUP_TIMENSEC" in STR):
        return 1
    else :
        return 0
#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def Get_MessagewithSignalName(SignalName , CAN_Database):
    STR_NAME_MESSAGE = None
    for mess in CAN_Database.messages :
        for signal in mess.signals :
            if (SignalName == signal.name) :
                STR_NAME_MESSAGE = mess.name
                return STR_NAME_MESSAGE
    return STR_NAME_MESSAGE
#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def GetPeriodicity(MessName , CAN_Database):
    STR_NAME_MESSAGE = "None"
    message = CAN_Database.get_message_by_name(MessName)
    return message.cycle_time

#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def LineSignalMessage(MessName , CAN_Database, DicoSigName):
    MessReq= MessName + "_REQ"
    Signame_Sig = DicoSigName+"_Sig"



#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def create_function_update_Message_Information(MessName_NAME,CAN_Database_NAME,Dico_Sig_Mess_NAME,CAPL_FILE_NAME) :
    CAPL_FILE_NAME.write("FUN_Update_Message_"+MessName_NAME+"()\n")
    CAPL_FILE_NAME.write("{\n")
    for Key_Sig in Dico_Sig_Mess_NAME.keys :
        if (MessName_NAME == Dico_Sig_Mess_NAME[Key_Sig] ):
            #this signal have to go to list message
            CAPL_FILE_NAME.write(LineSignalMessage(MessName_NAME , CAN_Database_NAME, Dico_Sig_Mess_NAME[Key_Sig]))
    CAPL_FILE_NAME.write("}\n")

#----------------------------------------------------------------------------
#------------------------------ Function ------------------------------------
#----------------------------------------------------------------------------
def Create_CAPL_VAR_ON_TIMER_EVENT(TabSignal , CAN_Database , CAPL_FILE_IN):
    try:
        CAPL_FILE = open(CAPL_FILE_IN ,'w')
    except IOError:
        print('cannot open', CAPL_FILE_IN)
        sys.exit(0)
    Dico_Sig_Mess={}
    for sigName in TabSignal :
        Mess_Name = Get_MessagewithSignalName(sigName,CAN_Database)
        if (Mess_Name is not None):
            Dico_Sig_Mess[sigName] = Mess_Name
    TabMessage = []
    for MessName in Dico_Sig_Mess.values() :
        if MessName not in TabMessage :
            TabMessage.append(MessName)
    for Message in TabMessage :
        CAPL_FILE.write("message "+ Message +" " + Message +"_REQ"+ "                      //"+str(GetPeriodicity(Message,CAN_Database))+"_ms periodicity\n")
#    for Message in TabMessage  :
#        create_function_update_Message_Information(Message,CAN_Database,Dico_Sig_Mess,CAPL_FILE)
    CAPL_FILE.close()

#----------------------------------------------------------------------------
#------------------------------Operation ------------------------------------
#----------------------------------------------------------------------------
#open DBC file
try:
    DBC_OPEN_FILE = open(DBC_PATH + DBC_FILE_NAME,'r')
except IOError:
        print('cannot open', DBC_FILE_NAME)
        sys.exit(0)
#due to error a new DBC file will be created without errors

try:
    DBC_OX_FILE = open(DBC_PATH + "OXFile"+ DBC_FILE_NAME ,'w')
except IOError:
        print('cannot open', "OXFile"+ DBC_FILE_NAME)
        sys.exit(0)
for Lines in DBC_OPEN_FILE.readlines() :
    if (check_Line_Prob(Lines) == 0) :
        DBC_OX_FILE.writelines(Treat(Lines))
DBC_OPEN_FILE.close()
DBC_OX_FILE.close()
#End Open and creation DBC file
#load DB_File
DB_database=cantools.database.load_file(DBC_PATH + "OXFile"+ DBC_FILE_NAME)

#Analyze Excel Sheet
Input_Signal_List = get_Input_Signal_List_From_ExcelSheet(ExcelSheet)
Create_File_Of_System_Variable(Input_Signal_List , DB_database, Sysvar_File)
Create_CAPL_VAR_ON_TIMER_EVENT(Input_Signal_List , DB_database, CAPL_File)
#Mess_Hash_Signal = get_Hash_Tab_From_Signal_List(Input_Signal_List , DB_database)




#for mess in DB_database.messages :
#
#    if ("FRONT_RADAR_GEN" not in mess.senders) :
#        print(mess)
    #example_message = DB_database.get_message_by_name(mess)
    #pprint(example_message.signals)
#print(DB_database.messages)
#----------------------------------------------------------------------------







