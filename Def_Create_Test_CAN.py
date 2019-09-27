#######################################################
# PSA
# date: 18/01/2019
# username:
# name: YJE
# description: Def create test can
#######################################################
#!/usr/bin/env python

#import Libs
import urllib.request as urllib2
import openpyxl
import os
import glob
from openpyxl import Workbook
######################################################

#From DOCINFO download the file Data_Conversion_Form_R5.xlsx

#Function section
######################################################
def GET_CPP_FILES(SRC_PATH,EXT) :
    TAB_FILES=[]
    for dossier, sous_dossiers, fichiers in os.walk(SRC_PATH):
        for Files in fichiers:
            File_Path_Name = os.path.join(dossier, Files)
            if (EXT.upper() in File_Path_Name.upper()):
                TAB_FILES.append(File_Path_Name)

                
    return TAB_FILES
        
######################################################
    
######################################################
def Read_XLSX_FILE(XLSX_FILE_PATH) :
    CAN_INT = []
    #open file
    try:
        XlFile= openpyxl.load_workbook(XLSX_FILE_PATH)
        #File_Open=open(XLSX_FILE_PATH,'r')
    except:
        print("file not opned")
    
    sheet_obj = XlFile.active
    m_row = sheet_obj.max_row
    #print(sheet_obj)
    #print(m_row)
    for i in range(3, m_row + 1):
        cell_obj_Number             = sheet_obj.cell(row = i, column = 1)
        cell_obj_CAN_SIG_NAME       = sheet_obj.cell(row = i, column = 2)
        cell_obj_INTERFACE_NAME     = sheet_obj.cell(row = i, column = 4)
        cell_obj_ARXML_DATA_NAME    = sheet_obj.cell(row = i, column = 5)
        cell_obj_INT_OUT_TYPE       = sheet_obj.cell(row = i, column = 6)
        #check line is not empty
        if (((cell_obj_CAN_SIG_NAME.value) is not None ) and ((cell_obj_INTERFACE_NAME.value) is not None) and ((cell_obj_ARXML_DATA_NAME.value) is not None) and ((cell_obj_INT_OUT_TYPE.value) is not None)) :
            CAN_INT.append((cell_obj_CAN_SIG_NAME.value)+","+(cell_obj_INTERFACE_NAME.value) +","+ (cell_obj_ARXML_DATA_NAME.value) +","+ (cell_obj_INT_OUT_TYPE.value))
            #print((cell_obj_CAN_SIG_NAME.value)+","+(cell_obj_INTERFACE_NAME.value) +","+ (cell_obj_ARXML_DATA_NAME.value) +","+ (cell_obj_INT_OUT_TYPE.value))
    # Write data to file
    return CAN_INT
        
######################################################


######################################################
def Write_Head_EXCEL_File(EXCEL_TEST_FILE_NAME ) :
    EXCEL_TEST_FILE_NAME['B16'] = "TEST PROCEDURE"

######################################################